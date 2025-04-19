import polars as pl
import os
from openpyxl import load_workbook
from xlsxwriter import Workbook
from typing import Dict, Union, List, Any

pl.Config.set_tbl_rows(100)
pl.Config.set_tbl_width_chars(9999)
pl.Config.set_fmt_str_lengths(100)

files = os.listdir('Reports')

def write_report_demontage(data: list | dict, workbook: Workbook):
    ws = workbook['демонтаж']
    start_row = 6

    # Если data - словарь, преобразуем в список из одного элемента
    if isinstance(data, dict):
        data = [data]

    for item in data:
        # Ищем первую полностью пустую строку
        while ws[f'A{start_row}'].value not in (None, ""):
            print(f"Строка {start_row} занята: {ws[f'A{start_row}'].value}")
            start_row += 1

        # Заполняем данные
        ws[f"A{start_row}"].value = item.get("sap")
        ws[f"B{start_row}"].value = item.get("name")
        ws[f"C{start_row}"].value = item.get("baseStation", item.get('warehouse'))  # Используем get с default
        ws[f"D{start_row}"].value = item.get("destination")
        # ws[f"E{start_row}"].value = item.get("type")

        print(f'Демонтаж внесен в строку {start_row}')
        start_row += 1

    # Вставляем пустую строку после последней записи
    ws.insert_rows(start_row)
    print(f"Добавлена пустая строка {start_row}")

def write_report_montage(data: list | dict, workbook: Workbook):

    ws = workbook['монтаж']
    start_row = 9

    # Унифицируем входные данные (работаем с list)
    if isinstance(data, dict):
        data = [data]


    for item in data:
        # Пропускаем записи не типа 'montage'
        if item.get('type') != 'montage':
            continue

        # Ищем первую пустую строку
        while ws[f'A{start_row}'].value not in (None, ""):
            start_row += 1

        # Заполняем данные
        ws[f"A{start_row}"].value = item.get("name")
        ws[f"B{start_row}"].value = 'монтаж'
        ws[f"C{start_row}"].value = item.get("destination", 'не указано')
        ws[f"D{start_row}"].value = item.get("count")
        ws[f"E{start_row}"].value = 'новое' if item.get("sap") == 'ТМЦ' else 'б/у'

        print(f'Монтаж внесен в строку {start_row}')
        start_row += 1

    # Вставляем пустую строку после последней записи
    if len(data) > 0:
        ws.insert_rows(start_row)
        print(f"Добавлена пустая строка {start_row}")

def new_report(data:str|dict, file_name = "Reports/template.xlsx"):
    '''

    :param data: словарь значений для вставки в СИМ/демонтаж (POST)
    :param path: путь к файлу СИМ/демонтаж (по умолчанию открывается пустой шаблон)
    :return: сохранение в excel
    '''



    print(os.getcwd())
    wb = load_workbook(file_name)
    write_report_montage(data=data, workbook=wb)
    write_report_demontage(data=data, workbook=wb)
    wb.save("v1_template.xlsx")
    wb.close()

def add_items(data:dict):

    rows = []

    for item in data['items']:
        row = {
            'id': item['id'],
            'type': item['type'],
            'destination': item.get('destination', None),  # Используем get на случай отсутствия ключа
            **item['data']  # Распаковываем содержимое data в тот же словарь
        }
        rows.append(row)

    return new_report(data=rows)

def combine_reports():

    combined_dem = pl.DataFrame()
    combined_mon = pl.DataFrame()

    for file in files:

        df_report_m = pl.read_excel('Reports/' + file, sheet_name='монтаж', read_options={"header_row": 6})
        df_report_m = df_report_m.filter(pl.col('БС') != 'null')
        df_report_m = df_report_m.filter(pl.col('Материал (новое/БУ)') == 'новое')
        combined_mon = pl.concat([combined_mon, df_report_m], how='vertical')

        df_report_d = pl.read_excel('Reports/' + file, sheet_name='демонтаж', read_options={"header_row": 4})
        df_report_d = df_report_d.filter(pl.col('NS___') != 'null')
        df_report_d = df_report_d.filter(pl.col('Перемещение осуществляется на склад/сайт').str.contains(r'(?i)NS'))
        combined_dem = pl.concat([combined_dem, df_report_d], how='vertical')

    return combined_dem, combined_mon

def read_reports(
        file_names: Union[str, List[str]] = None,
        template: str = 'template.xlsx',
        as_dataframes: bool = True
) -> Dict[str, Dict[str, Union[List[Dict[str, Any]], pl.DataFrame]]]:
    """
    Загрузка отчетов СИМ из одного или нескольких файлов Excel

    Параметры:
        file_names: имя файла (str) или список файлов (List[str]). Если None, используется template.
        template: имя шаблонного файла (используется, если file_names не указан)
        as_dataframes: если True - возвращает DataFrame, если False - преобразует в список словарей

    Возвращает:
        Словарь, где:
        - ключ: имя файла (str)
        - значение: словарь с ключами 'монтаж' и 'демонтаж', содержащие:
            * DataFrame (если as_dataframes=True)
            * список словарей (если as_dataframes=False)
    """
    # Обработка входных параметров
    if file_names is None:
        files = [template]
    elif isinstance(file_names, str):
        files = [file_names]
    else:
        files = file_names

    result = {}

    for file in files:
        file_path = f'Reports/{file}'

        try:
            # Обработка листа 'монтаж'
            df_montage = (
                pl.read_excel(
                    file_path,
                    sheet_name='монтаж',
                    read_options={"header_row": 6},
                    infer_schema_length=0
                )
                .filter(pl.col('БС') != 'null')
                .fill_null('')
            )

            # Обработка листа 'демонтаж'
            df_demontage = (
                pl.read_excel(
                    file_path,
                    sheet_name='демонтаж',
                    read_options={"header_row": 4}
                )
                .filter(pl.col('NS___') != 'null')
                .fill_null('')
            )

            # Преобразуем в словари если требуется
            if not as_dataframes:
                df_montage = df_montage.to_dicts()
                df_demontage = df_demontage.to_dicts()

            result[file] = {
                'монтаж': df_montage,
                'демонтаж': df_demontage
            }

        except Exception as e:
            print(f"Ошибка при обработке файла {file}: {e}")
            result[file] = {
                'монтаж': [] if not as_dataframes else pl.DataFrame(),
                'демонтаж': [] if not as_dataframes else pl.DataFrame()
            }

    return result


# data_source = {'items': [{'id': '140000064378-0', 'type': 'demontage', 'destination': 'Не выбрано', 'data': {}}, {'id': '140000000869-1', 'type': 'demontage', 'destination': 'Не выбрано', 'data': {'name': 'Блок модема ММU2 MMU2 B ROJ2081301/10', 'count': 1, 'baseStation': 'NS001152', 'destination': 'Не выбрано', 'sap': '140000000869'}}, {'id': 'P-T223046.54.9996-6', 'type': 'montage', 'data': {}}, {'id': 'P-T225003.54.0573-681', 'type': 'montage', 'data': {'name': 'Аккумуляторная батарея\xa0 FTS 12-150 X', 'sppElement': 'P-T225003.54.0573', 'count': 1, 'warehouse': 'K026', 'destination': 'NS001152', 'party': '0002108828', 'sap': 'ТМЦ'}}]}
# data =[{'id': '140000071873-34', 'type': 'demontage', 'destination': 'Не выбрано', 'name': 'Приемопередающий модуль FRMF 6TX800 360W', 'count': 1, 'baseStation': 'NS001588', 'sap': '140000071873'}, {'id': 'P-T221001.54.9996-639', 'type': 'montage', 'destination': 'KZ01', 'name': 'Приемопередающий модуль FRGX RFM 3 2100', 'sppElement': 'P-T221001.54.9996', 'count': 1, 'warehouse': 'KZ01', 'party': 'Z000104899', 'sap': '140000031564'}]
# add_items(data=data_source)
# print()

# print(read_reports(file=files[1]))

def upload_reports():
    # file_name = request.get_json()
    report = read_reports(files[1:2], as_dataframes=False)

    return report

# print(upload_reports())